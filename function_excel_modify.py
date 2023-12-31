import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pymystem3 import Mystem
import nltk
import pymorphy3
import sys

nltk.download('stopwords')

# Инициализируем MorphAnalyzer
morph = pymorphy3.MorphAnalyzer()

# Функция для определения пола по имени
def determine_gender(name_parts, mystem):
    for word in name_parts.split():
        analysis = mystem.analyze(word)
        for part in analysis:
            if 'analysis' in part and part['analysis']:
                gr = part['analysis'][0]['gr']
                if 'имя' in gr and 'жен' in gr:
                    return 'female'
                elif 'имя' in gr and 'муж' in gr:
                    return 'male'
    return 'unknown'

# Функция для извлечения имени из строки
def extract_name(name_str, mystem):
    analysis = mystem.analyze(name_str)
    name_parts = []
    for word in analysis:
        if 'analysis' in word and word['analysis']:
            gr = word['analysis'][0]['gr']
            if 'имя' in gr or 'отч' in gr:
                name_parts.append(word['text'])
    return ' '.join(name_parts) if name_parts else 'Не определено'

# Функция для склонения должности в дательный падеж
def inflect_to_dative(position, morph):
    words = position.split()
    inflected_words = []
    
    for i, word in enumerate(words):
        parsed_word = morph.parse(word)[0]  # Получаем наиболее вероятный разбор слова
        
        # Если слово - предлог, союз и т.п., не склоняем его
        if {'PREP', 'CONJ', 'PRCL'}.intersection(set(parsed_word.tag.grammemes)):
            inflected_word = word
        else:
            # Склоняем слово, если это возможно
            inflected_word = parsed_word.inflect({'datv'})
            inflected_word = inflected_word.word if inflected_word else word
        
        # Добавляем слово в результат
        inflected_words.append(inflected_word)
    
    # Преобразуем первое слово в строке с заглавной буквы
    if inflected_words:
        inflected_words[0] = inflected_words[0].capitalize()
    
    return ' '.join(inflected_words)

def process_excel_file(file_path):
    # Инициализация Mystem
    mystem = Mystem()

    # Загружаем книгу Excel и активный лист
    book = load_workbook(file_path)
    sheet = book.active

    # Преобразовываем активный лист Excel в DataFrame
    data = list(sheet.values)
    columns = next(iter(data))
    df = pd.DataFrame(data[1:], columns=columns)

    for index in range(len(df)):
        # Получаем строку как Series для изменения
        row = df.iloc[index]
        
        print(index)
        # Извлекаем имя
        name = extract_name(row['Имя адресата'], mystem)
        # Определяем пол
        gender = determine_gender(name, mystem)

        # Обновляем 'Имя для рассылки' и 'Обращение' в DataFrame
        df.at[index, 'Имя для рассылки'] = name
        if len(name.split()) == 1:
            greeting = f"{name}, здравствуйте!"
        else:
            greeting = f"Уважаем{'ая' if gender == 'female' else 'ый'} {name},"
        df.at[index, 'Обращение'] = greeting

        # Склоняем 'Должность' в дательный падеж и обновляем в DataFrame
        position = row['Должность']
        df.at[index, 'Должность Дательный падеж'] = inflect_to_dative(position, morph)

  
    # Сохраняем изменения, не трогая остальные ячейки
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    book.save(file_path)

# Вызываем функцию с путем к вашему файлу
if __name__ == '__main__':
    # Проверяем, передан ли путь к файлу в качестве аргумента командной строки
    if len(sys.argv) > 1:
        file_path = sys.argv[1]  # Используем имя файла, переданное из .bat файла
        process_excel_file(file_path)
    else:
        file_path = 'list.xlsx'
        process_excel_file(file_path)        
        #print("Ошибка: Не указан путь к файлу Excel.")
        #sys.exit(1)
